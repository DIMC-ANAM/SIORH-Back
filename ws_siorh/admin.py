import base64
from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from .models import PlantillaFin, BajasFin, Movimiento, Acceso2025
from django.db.models import Count, Q
import openpyxl
from openpyxl.drawing.image import Image
from io import BytesIO
from collections import defaultdict
from datetime import datetime
from django.utils import timezone
from django.shortcuts import redirect
from django.contrib import messages
from django.urls import path, reverse

# Importamos lo necesario para la bitácora
try:
    from auditoria.models import RegistroActividad
except Exception:
    class _DummyRegistroManager:
        @staticmethod
        def create(*args, **kwargs):
            return None

    class RegistroActividad:
        objects = _DummyRegistroManager()

try:
    from auditoria.utils import log_user_view  # Función de ayuda para registrar
except Exception:
    def log_user_view(*args, **kwargs):
        return None

# --- "FÁBRICA" DE FILTROS ---
def crear_filtro_conteo(nombre_campo, titulo_filtro):
    class CustomConteoFiltro(admin.SimpleListFilter):
        title = titulo_filtro
        parameter_name = nombre_campo
        def lookups(self, request, model_admin):
            qs = model_admin.get_queryset(request)
            counts = qs.values(self.parameter_name).annotate(count=Count('pk')).order_by('-count')
            return [(item[self.parameter_name], f"{item[self.parameter_name]} ({item['count']})") for item in counts if item[self.parameter_name] is not None]
        def queryset(self, request, queryset):
            if self.value():
                return queryset.filter(**{self.parameter_name: self.value()})
            return queryset
    return CustomConteoFiltro

# --- CLASE DE ADMIN PARA PlantillaFin ---
class PlantillaFinAdmin(admin.ModelAdmin):
    actions = ['exportar_a_excel_con_imagenes']

    list_display = (
        'colored_miniatura', 'colored_nombres', 'colored_rfc', 'colored_num_empleado',
        'colored_posicion', 'colored_nivel', 'colored_programa', 'colored_dg_o_aduana_compactada',
        'colored_unidad_administrativa', 'colored_departamento', 'colored_nombre_puesto_funcional',
        'colored_estado_nomina'
    )
    search_fields = ('nombres', 'estado_nomina', 'posicion','nivel', 'num_empleado', 'rfc','nombre_puesto_funcional','dg_o_aduana_compactada','programa')
    fieldsets = (
        ('Información Personal', { 'fields': ('mostrar_foto_de_base_de_datos','nombres', 'rfc', 'curp', 'num_empleado') }),
        ('Detalles del Puesto', { 'fields': ('posicion', 'estado_nomina', 'nivel', 'codigo_presupuestal', 'nombre_puesto_funcional', 'tipo_de_contratacion') }),
        ('Control de Asistencia', {
            'classes': ('collapse', 'tab', 'empty-form'),
            'fields': ('calendario_asistencia',),
        }),
    )
    readonly_fields = ('mostrar_foto_de_base_de_datos', 'calendario_asistencia')
    
    # --- MÉTODOS PARA COLOREAR Y MOSTRAR FOTOS ---
    def _get_row_color(self, obj):
        if obj.programa == 'Sin Código' or obj.estado_nomina == 'Baja': return '#FFDDDD'
        elif obj.estado_nomina == 'Vacante' and 'H00' in str(obj.programa): return '#D4EDDA'
        return ''

    @admin.display(description='Foto')
    def colored_miniatura(self, obj):
        color = self._get_row_color(obj)
        content = self.mostrar_miniatura(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:1px; text-align:center;">{}</div>', color, content)

    @admin.display(description='Nombre Completo', ordering='nombres')
    def colored_nombres(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.nombres)

    @admin.display(description='RFC', ordering='rfc')
    def colored_rfc(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.rfc)

    @admin.display(description='Num Empleado', ordering='num_empleado')
    def colored_num_empleado(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.num_empleado)

    @admin.display(description='Posición', ordering='posicion')
    def colored_posicion(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.posicion)

    @admin.display(description='Nivel', ordering='nivel')
    def colored_nivel(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.nivel)

    @admin.display(description='Programa', ordering='programa')
    def colored_programa(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.programa)
        
    @admin.display(description='DG o Aduana', ordering='dg_o_aduana_compactada')
    def colored_dg_o_aduana_compactada(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.dg_o_aduana_compactada)

    @admin.display(description='Unidad Admin.', ordering='unidad_administrativa')
    def colored_unidad_administrativa(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.unidad_administrativa)

    @admin.display(description='Departamento', ordering='departamento')
    def colored_departamento(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.departamento)

    @admin.display(description='Puesto Funcional', ordering='nombre_puesto_funcional')
    def colored_nombre_puesto_funcional(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.nombre_puesto_funcional)

    @admin.display(description='Estado Nómina', ordering='estado_nomina')
    def colored_estado_nomina(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.estado_nomina)

    def mostrar_foto_de_base_de_datos(self, obj):
        if obj.foto:
            imagen_base64 = base64.b64encode(obj.foto).decode('utf-8')
            return format_html(f'<img src="data:image/jpeg;base64,{imagen_base64}" style="max-width: 200px; height: auto;" />')
        return "No hay imagen en la base de datos."
    mostrar_foto_de_base_de_datos.short_description = 'Fotografía'

    def mostrar_miniatura(self, obj):
        if obj.foto:
            imagen_base64 = base64.b64encode(obj.foto).decode('utf-8')
            return format_html(f'<img src="data:image/jpeg;base64,{imagen_base64}" style="width: 60px; height: auto;" />')
        return "N/A"

    @admin.display(description='Calendario de asistencia')
    def calendario_asistencia(self, obj):
        if not obj:
            return ""

        num_empleado = getattr(obj, "num_empleado", None)
        if not num_empleado:
            return mark_safe("No se encontró número de empleado para enlazar accesos.")

        # Normalizar empleado: con y sin ceros a la izquierda
        empleado_raw = str(num_empleado).strip()
        empleado_sin_ceros = empleado_raw.lstrip('0') or empleado_raw

        # 1) Traer los accesos del empleado (de cualquier año)
        accesos = Acceso2025.objects.filter(
            Q(empleado=empleado_raw) | Q(empleado=empleado_sin_ceros)
        )

        if not accesos:
            return mark_safe("No se encontraron registros de accesos para este empleado.")

        # Helpers de nombres en español
        DIAS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        MESES_ES = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]

        # 2) Agrupar por fecha -> lista de (hora, temp)
        dias = defaultdict(list)  # { '2025-02-16': [ {'hora': '07:55:00', 'temp': 36.4, 'dt': datetime}, ... ] }
        fechas_dt = []  # para saber qué meses/años hay

        for acc in accesos:
            raw = acc.date_raw or ""
            if len(raw) < 19:
                continue

            fecha = raw[:10]         # 'YYYY-MM-DD'
            hora = raw[11:19]        # 'HH:MM:SS' (desde posición 11, formato: "YYYY-MM-DD HH:MM:SS")

            try:
                dt = datetime.strptime(fecha, "%Y-%m-%d")
                fechas_dt.append(dt)
            except Exception:
                continue

            # Temperatura: extraer número decimal de param0 (ej: '36.7°C, ...')
            temp_val = None
            if acc.param0:
                # Buscar patrón: número con opcional punto/coma decimal
                import re
                match = re.search(r'(\d+)[.,](\d+)', acc.param0)
                if match:
                    try:
                        temp_val = float(f"{match.group(1)}.{match.group(2)}")
                    except Exception:
                        temp_val = None

            dias[fecha].append({
                "hora": hora,    # 'HH:MM:SS'
                "temp": temp_val,
                "dt": dt,
            })

        if not dias:
            return mark_safe("No se pudieron procesar registros de accesos válidos para este empleado.")

        # Helpers para promedios
        def hora_a_segundos(h):
            # h: 'HH:MM' o 'HH:MM:SS'
            if not h or h == "-":
                return None
            partes = h.split(":")
            if len(partes) < 2:
                return None
            try:
                hh = int(partes[0])
                mm = int(partes[1])
                ss = int(partes[2]) if len(partes) > 2 else 0
                return hh * 3600 + mm * 60 + ss
            except Exception:
                return None

        def segundos_a_hhmm(seg):
            if seg is None:
                return "-"
            hh = seg // 3600
            mm = (seg % 3600) // 60
            return f"{hh:02d}:{mm:02d}"

        # 3) Calcular entrada/salida por día y agrupar por mes
        meses = defaultdict(list)  # { '2025-02': [ { 'dt':..., 'fecha_label':..., 'entrada_hora':..., ... }, ... ] }

        for fecha_str, registros in dias.items():
            dt = registros[0]["dt"]  # todos los del día comparten fecha

            # separa en mañana (<12) y tarde (>=12)
            registros_manan = []
            registros_tarde = []

            for r in registros:
                hora_txt = r["hora"]
                try:
                    hh = int(hora_txt[:2])
                except Exception:
                    hh = 0
                if hh < 12:
                    registros_manan.append(r)
                else:
                    registros_tarde.append(r)

            # Entrada = menor hora de la mañana (si hay)
            entrada = min(registros_manan, key=lambda r: r["hora"]) if registros_manan else None
            # Salida = mayor hora de la tarde (si hay)
            salida = max(registros_tarde, key=lambda r: r["hora"]) if registros_tarde else None

            # Etiqueta tipo: "Lunes 3 Noviembre 2025"
            weekday_idx = dt.weekday()  # 0 = lunes
            dia_nombre = DIAS_ES[weekday_idx]
            mes_nombre = MESES_ES[dt.month - 1]
            etiqueta_dia = f"{dia_nombre} {dt.day} {mes_nombre} {dt.year}"

            clave_mes = dt.strftime("%Y-%m")     # '2025-11'
            nombre_mes = f"{mes_nombre} {dt.year}"

            # formato 'HH:MM' o '-'
            if entrada:
                entrada_hora_str = entrada["hora"][:5]
                entrada_temp_val = entrada["temp"]
            else:
                entrada_hora_str = "-"
                entrada_temp_val = None

            if salida:
                salida_hora_str = salida["hora"][:5]
                salida_temp_val = salida["temp"]
            else:
                salida_hora_str = "-"
                salida_temp_val = None

            meses[clave_mes].append({
                "dt": dt,
                "fecha_label": etiqueta_dia,
                "entrada_hora": entrada_hora_str,
                "entrada_temp": entrada_temp_val,
                "salida_hora": salida_hora_str,
                "salida_temp": salida_temp_val,
                "nombre_mes": nombre_mes,
            })

        # 4) Ordenar meses y decidir cuál es el "vigente" (mes actual si tiene datos)
        meses_ordenados = dict(sorted(meses.items(), key=lambda x: x[0]))  # orden cronológico

        hoy = timezone.localtime(timezone.now()).date()
        clave_mes_hoy = hoy.strftime("%Y-%m")  # ej. '2026-02'

        if clave_mes_hoy in meses_ordenados:
            vigente_clave_mes = clave_mes_hoy
        elif fechas_dt:
            dt_max = max(fechas_dt)
            vigente_clave_mes = dt_max.strftime("%Y-%m")
        else:
            vigente_clave_mes = next(iter(meses_ordenados.keys()))

        # 5) Construir HTML (nav + tablas por mes + estadísticas por mes)
        html = """
        <style>
            .asistencia-nav {
                margin: 8px 0 12px 0;
                font-size: 11px;
            }
            .asistencia-nav a {
                margin-right: 6px;
                text-decoration: none;
                padding: 3px 6px;
                border-radius: 3px;
                border: 1px solid #ccc;
                color: #333;
            }
            .asistencia-nav a.asistencia-nav-activo {
                background-color: #800000;
                color: #fff;
                border-color: #800000;
            }
            .asistencia-mes-title {
                margin-top: 10px;
                font-weight: bold;
                font-size: 13px;
                color: #800000;
            }
            .asistencia-tabla {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 10px;
                font-size: 11px;
            }
            .asistencia-tabla th, .asistencia-tabla td {
                border: 1px solid #ddd;
                padding: 4px 6px;
                text-align: center;
            }
            .asistencia-tabla th {
                background-color: #f5f5f5;
                font-weight: bold;
            }
            .asistencia-fila-par {
                background-color: #fafafa;
            }
            .asistencia-fila-impar {
                background-color: #ffffff;
            }
            .asistencia-mes-wrapper {
                display: none;
            }
            .asistencia-mes-wrapper.activo {
                display: block;
            }
            .asistencia-resumen-mes {
                font-size: 11px;
                margin: 4px 0 8px 0;
                padding: 4px 6px;
                background-color: #f0f0f0;
                border-radius: 4px;
            }
        </style>
        """

        # Navegación por meses (tabs)
        html += '<div class="asistencia-nav">'
        claves_mes_lista = list(meses_ordenados.keys())
        for clave_mes in claves_mes_lista:
            lista_dias = meses_ordenados[clave_mes]
            if not lista_dias:
                continue
            nombre_mes = lista_dias[0]["nombre_mes"]
            activo_cls = "asistencia-nav-activo" if clave_mes == vigente_clave_mes else ""
            html += f'<a href="#!" class="asistencia-nav-link {activo_cls}" data-mes="{clave_mes}">{nombre_mes}</a>'
        html += '</div>'

        # Contenedor de meses
        html += '<div id="asistencia-meses-contenedor">'

        for clave_mes, lista_dias in meses_ordenados.items():
            if not lista_dias:
                continue

            nombre_mes = lista_dias[0]["nombre_mes"]
            activo_cls = "activo" if clave_mes == vigente_clave_mes else ""

            # --- Estadísticas por mes ---
            entrada_segundos_mes = []
            salida_segundos_mes = []
            for d in lista_dias:
                ent_seg = hora_a_segundos(d["entrada_hora"])
                if ent_seg is not None:
                    entrada_segundos_mes.append(ent_seg)

                # solo contamos salida si realmente hay salida (no "-")
                if d["salida_hora"] and d["salida_hora"] != "-":
                    sal_seg = hora_a_segundos(d["salida_hora"])
                    if sal_seg is not None:
                        salida_segundos_mes.append(sal_seg)

            prom_entrada_mes = segundos_a_hhmm(
                sum(entrada_segundos_mes) // len(entrada_segundos_mes)
            ) if entrada_segundos_mes else "-"
            prom_salida_mes = segundos_a_hhmm(
                sum(salida_segundos_mes) // len(salida_segundos_mes)
            ) if salida_segundos_mes else "-"

            # --- Wrapper del mes ---
            html += f'<div class="asistencia-mes-wrapper {activo_cls}" id="asistencia-mes-{clave_mes}">'
            html += f'<div class="asistencia-mes-title">{nombre_mes}</div>'

            # Estadísticas mensuales encima de la tabla
            html += f"""
            <div class="asistencia-resumen-mes">
                <strong>Estadísticas del mes:</strong>
                Promedio de entrada: <strong>{prom_entrada_mes}</strong>,
                Promedio de salida: <strong>{prom_salida_mes}</strong>
            </div>
            """

            html += """
            <table class="asistencia-tabla">
                <thead>
                    <tr>
                        <th>Día</th>
                        <th>Hora Entrada</th>
                        <th>Temp. Entrada</th>
                        <th>Hora Salida</th>
                        <th>Temp. Salida</th>
                    </tr>
                </thead>
                <tbody>
            """

            # ordenar días del mes por fecha real
            lista_dias_ordenados = sorted(lista_dias, key=lambda d: d["dt"])

            for idx, dia in enumerate(lista_dias_ordenados):
                fila_clase = "asistencia-fila-par" if idx % 2 == 0 else "asistencia-fila-impar"

                # Formateo de temperaturas
                ent_temp_val = dia["entrada_temp"]
                sal_temp_val = dia["salida_temp"]

                if isinstance(ent_temp_val, (int, float, complex)):
                    temp_ent = f"{float(ent_temp_val):.1f}°"
                elif ent_temp_val:
                    temp_ent = f"{ent_temp_val}°"
                else:
                    temp_ent = "-"

                if isinstance(sal_temp_val, (int, float, complex)):
                    temp_sal = f"{float(sal_temp_val):.1f}°"
                elif sal_temp_val:
                    temp_sal = f"{sal_temp_val}°"
                else:
                    temp_sal = "-"

                html += f"""
                    <tr class="{fila_clase}">
                        <td>{dia['fecha_label']}</td>
                        <td>{dia['entrada_hora']}</td>
                        <td>{temp_ent}</td>
                        <td>{dia['salida_hora']}</td>
                        <td>{temp_sal}</td>
                    </tr>
                """

            html += """
                </tbody>
            </table>
            </div>
            """

        html += "</div>"

        # JS mínimo para cambiar de mes (y por tanto de estadísticas mensuales)
        html += """
        <script>
            (function() {
                var links = document.querySelectorAll('.asistencia-nav-link');
                if (!links.length) return;

                links.forEach(function(link) {
                    link.addEventListener('click', function(e) {
                        e.preventDefault();
                        var mes = this.getAttribute('data-mes');

                        // activar/desactivar links
                        links.forEach(function(l2) {
                            l2.classList.remove('asistencia-nav-activo');
                        });
                        this.classList.add('asistencia-nav-activo');

                        // mostrar/ocultar contenedores de meses
                        var wrappers = document.querySelectorAll('.asistencia-mes-wrapper');
                        wrappers.forEach(function(w) {
                            w.classList.remove('activo');
                        });
                        var activo = document.getElementById('asistencia-mes-' + mes);
                        if (activo) {
                            activo.classList.add('activo');
                        }
                    });
                });
            })();
        </script>
        """

        return mark_safe(html)

    # --- ACCIÓN DE EXPORTAR CON REGISTRO EN BITÁCORA ---
    @admin.action(description="Exportar a Excel con Imágenes")
    def exportar_a_excel_con_imagenes(self, request, queryset):
        RegistroActividad.objects.create(usuario=request.user, tipo_accion='export', detalles=f"Exportó a Excel {queryset.count()} registros de Plantilla de Personal.")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Plantilla de Personal"
        headers = ['Posición', 'Nombre', 'RFC', 'SMN', 'Foto']
        sheet.append(headers)
        for registro in queryset:
            if registro.foto:
                try:
                    image_stream = BytesIO(registro.foto)
                    img = Image(image_stream)
                    proporcion = img.width / img.height
                    img.height = 80
                    img.width = 80 * proporcion
                except Exception: img = None
            else: img = None
            row_data = [registro.posicion, registro.nombres, registro.rfc, getattr(registro, 'smn', ''), '']
            sheet.append(row_data)
            if img:
                row_num = sheet.max_row
                sheet.add_image(img, f'E{row_num}')
                sheet.row_dimensions[row_num].height = 65
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="plantilla_personal.xlsx"'
        workbook.save(response)
        return response

    # --- REGISTRO DE CONSULTAS Y BÚSQUEDAS ---
    def changelist_view(self, request, extra_context=None):
        search_term = request.GET.get('q', '') 
        if search_term:
            log_user_view(request, f"Buscó '{search_term}' en Plantilla de Personal.", include_filters=False) # No incluir filtros si es búsqueda
        else:
            log_user_view(request, "Consultó la lista de Plantilla de Personal.", include_filters=True) # Incluir filtros si es consulta normal
        return super().changelist_view(request, extra_context)

    # --- REGISTRO DE VISTA DE DETALLE ---
    def change_view(self, request, object_id, form_url='', extra_context=None):
        try:
            obj = self.get_object(request, object_id)
            log_user_view(request, f"Consultó el detalle de Plantilla: {obj}", include_filters=False) # No incluir filtros aquí
        except Exception: 
            log_user_view(request, f"Intentó consultar detalle de Plantilla ID: {object_id}", include_filters=False)
        return super().change_view(request, object_id, form_url, extra_context)
        
    # --- GESTIÓN DE ACCIONES Y PERMISOS ---
    def get_actions(self, request):
        actions = super().get_actions(request)
        if 'delete_selected' in actions: del actions['delete_selected']
        return actions

    def has_delete_permission(self, request, obj=None):
        return False

    class Media:
        css = { 'all': ('core/css/admin_sticky_headers.css',) }

# --- CLASE DE ADMIN PARA BajasFin ---
class BajasFinAdmin(admin.ModelAdmin):
    actions = ['exportar_a_excel_con_imagenes']
    list_display = ('mostrar_miniatura','nombre', 'rfc','id_empleado', 'posicion','nivel','motivo', 'fecha_efectiva_personal', 'puesto_funcional','unidad_administrativa')
    search_fields = ('nombre', 'id_empleado', 'rfc', 'motivo','puesto_funcional','nivel','unidad_administrativa')
    list_filter = (crear_filtro_conteo('motivo', 'Motivo'), crear_filtro_conteo('nivel', 'Nivel'), crear_filtro_conteo('unidad_administrativa', 'Unidad Administrativa'), crear_filtro_conteo('posicion', 'Posición'))
    fieldsets = (('Información del Empleado', {'fields': ('nombre', 'id_empleado', 'rfc', 'curp')}), ('Detalles de la Baja', {'fields': ('motivo', 'estado_nomina', 'fecha_efectiva_personal', 'fecha_de_captura')}), ('Información del Puesto', {'fields': ('posicion', 'puesto_funcional', 'unidad_administrativa')}), ('Fotografía', {'fields': ('mostrar_foto_de_base_de_datos',)}))
    readonly_fields = ('mostrar_foto_de_base_de_datos',)
    
    # --- MÉTODOS PARA FOTOS ---
    def mostrar_foto_de_base_de_datos(self, obj):
        if obj.foto:
            imagen_base64 = base64.b64encode(obj.foto).decode('utf-8')
            return format_html(f'<img src="data:image/jpeg;base64,{imagen_base64}" style="max-width: 200px; height: auto;" />')
        return "No hay imagen en la base de datos."
    mostrar_foto_de_base_de_datos.short_description = 'Fotografía'

    def mostrar_miniatura(self, obj):
        if obj.foto:
            imagen_base64 = base64.b64encode(obj.foto).decode('utf-8')
            return format_html(f'<img src="data:image/jpeg;base64,{imagen_base64}" style="width: 60px; height: auto;" />')
        return "N/A"
    mostrar_miniatura.short_description = 'Foto'

    # --- ACCIÓN DE EXPORTAR CON REGISTRO EN BITÁCORA ---
    @admin.action(description="Exportar Bajas seleccionadas a Excel con imágenes")
    def exportar_a_excel_con_imagenes(self, request, queryset):
        RegistroActividad.objects.create(usuario=request.user, tipo_accion='export', detalles=f"Exportó a Excel {queryset.count()} registros de Bajas de Personal.")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Bajas de Personal"
        headers = ['ID Empleado', 'Nombre', 'RFC', 'Posición', 'Motivo', 'Fecha Efectiva', 'Foto']
        sheet.append(headers)
        for registro in queryset:
            if registro.foto:
                try:
                    image_stream = BytesIO(registro.foto)
                    img = Image(image_stream)
                    proporcion = img.width / img.height
                    img.height = 80
                    img.width = 80 * proporcion
                except Exception: img = None
            else: img = None
            row_data = [registro.id_empleado, registro.nombre, registro.rfc, registro.posicion, registro.motivo, registro.fecha_efectiva_personal, '']
            sheet.append(row_data)
            if img:
                row_num = sheet.max_row
                sheet.add_image(img, f'G{row_num}')
                sheet.row_dimensions[row_num].height = 65
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="bajas_personal.xlsx"'
        workbook.save(response)
        return response

    # --- REGISTRO DE CONSULTAS Y BÚSQUEDAS ---
    def changelist_view(self, request, extra_context=None):
        search_term = request.GET.get('q', '')
        if search_term:
            log_user_view(request, f"Buscó '{search_term}' en Bajas de Personal.", include_filters=False)
        else:
            log_user_view(request, "Consultó la lista de Bajas de Personal.", include_filters=True)
        return super().changelist_view(request, extra_context)

    # --- REGISTRO DE VISTA DE DETALLE ---
    def change_view(self, request, object_id, form_url='', extra_context=None):
        try:
            obj = self.get_object(request, object_id)
            log_user_view(request, f"Consultó el detalle de Baja: {obj}", include_filters=False)
        except Exception:
            log_user_view(request, f"Intentó consultar detalle de Baja ID: {object_id}", include_filters=False)
        return super().change_view(request, object_id, form_url, extra_context)

    # --- GESTIÓN DE ACCIONES ---
    def get_actions(self, request):
        actions = super().get_actions(request)
        if 'delete_selected' in actions: del actions['delete_selected']
        return actions


# --- REGISTRO DE MODELOS EN EL ADMIN ---
admin.site.register(PlantillaFin, PlantillaFinAdmin)
admin.site.register(BajasFin, BajasFinAdmin)
