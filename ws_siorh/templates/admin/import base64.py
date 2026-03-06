import base64
from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
from django.utils.safestring import mark_safe
# Importación CRÍTICA del nuevo modelo MovPos
from .models import PlantillaFin, BajasFin, CalendarioNomina2025, MovPos, Movimiento, Acceso2025, DocumentoExpediente
from django.db.models import Count
import openpyxl
from openpyxl.drawing.image import Image
from io import BytesIO

# Importamos lo necesario para la bitácora
from auditoria.models import RegistroActividad
from auditoria.utils import log_user_view # Función de ayuda para registrar

from datetime import datetime
from collections import defaultdict
from django.utils import timezone
from django.db.models import Q

from django.shortcuts import redirect
from django.contrib import messages
from django.urls import path, reverse

# --- "FÁBRICA" DE FILTROS (SIN CAMBIOS) ---
def crear_filtro_conteo(nombre_campo, titulo_filtro):
    class CustomConteoFiltro(admin.SimpleListFilter):
        title = titulo_filtro
        parameter_name = nombre_campo
        def lookups(self, request, model_admin):
            qs = model_admin.get_queryset(request)
            counts = qs.values(self.parameter_name).annotate(count=Count('pk')).order_by('-count')
            return [(item[self.parameter_name], f"{item[self.parameter_name]} ({item['count']})") for item in counts if item[self.parameter_name] is not None]
        def queryset(self, request, queryset):
            if self.value():
                return queryset.filter(**{self.parameter_name: self.value()})
            return queryset
    return CustomConteoFiltro

# ================================
#   INLINES DE DOCUMENTOS
# ================================
class DocumentoExpedienteActivoInline(admin.TabularInline):
    """
    Inline para ver/cargar documentos de expediente
    ligados a empleados ACTIVOS (PlantillaFin).
    """
    model = DocumentoExpediente
    fk_name = "empleado_activo"  # clave foránea que se usará en este inline
    extra = 0
    fields = (
        "tipo_documento", "archivo"
    )
    readonly_fields = ("estado_proceso", "fecha_subida","nombre_original",
        "anio_ingreso_anam",
        "estado_proceso",
        "fecha_subida")
    show_change_link = True

class DocumentoExpedienteBajaInline(admin.TabularInline):
    """
    Inline para ver/cargar documentos de expediente
    ligados a empleados dados de BAJA (BajasFin).
    """
    model = DocumentoExpediente
    fk_name = "empleado_baja"
    extra = 0
    fields = (
        "tipo_documento","archivo"
    )
    readonly_fields = ("estado_proceso", "fecha_subida","nombre_original",
        "anio_ingreso_anam",
        "estado_proceso",
        "fecha_subida")
    show_change_link = True

# -------------------------------------------------------------
# --- CLASE DE ADMIN PARA PlantillaFin (CORREGIDA) ---
# -------------------------------------------------------------
class PlantillaFinAdmin(admin.ModelAdmin):
    actions = ['exportar_a_excel_con_imagenes']
    # ... (resto de la clase PlantillaFinAdmin) ...

    inlines = [DocumentoExpedienteActivoInline]
    
    list_display = (
        'colored_miniatura',"generar_cedula_link","expediente_link", 'colored_nombres', 'colored_rfc', 'colored_num_empleado',
        'colored_posicion', 'colored_nivel','colored_smn', 'colored_programa', 'colored_dg_o_aduana_compactada',
        'colored_unidad_administrativa', 'colored_departamento', 'colored_nombre_puesto_funcional',
        'colored_estado_nomina','colored_tipo_de_personal_sedena_semar','colored_rango'
    )
    
    search_fields = ('nombres', 'estado_nomina', 'posicion','nivel', 'num_empleado', 'rfc','nombre_puesto_funcional','dg_o_aduana_compactada','programa')
    
    # 1. FIELDSETS: AÑADIMOS EL CAMPO DE HISTORIAL
    fieldsets = (
        ('Información Personal', { 'fields': ('mostrar_foto_de_base_de_datos','nombres', 'rfc', 'curp', 'num_empleado') }),
        # Asegúrate de que los fieldsets de pestañas apunten a los campos existentes
        ('Detalles del Puesto', {
            'classes': ('collapse', 'tab', 'empty-form'),
            'fields': ('posicion', 'estado_nomina', 'nivel', 'smn','codigo_presupuestal', 'nombre_puesto_funcional', 'tipo_de_contratacion')
        }),
        ('Ubicación Administrativa', {
            'classes': ('collapse', 'tab', 'empty-form'),
            'fields': ('dg_o_aduana_compactada', 'unidad_administrativa', 'departamento','tipo_de_personal_sedena_semar','rango') 
        }),
        # CRÍTICO: El campo que se genera con el método
        ('Histórico de Plaza', {
            'classes': ('collapse', 'tab', 'empty-form'), 
            'fields': ('historial_posiciones_detalle',) 
        }),
        
        ('Histórico de Movimientos de Personal', {
        'classes': ('collapse', 'tab', 'empty-form'),
        'fields': ('historico_movimientos_personal',),
        }),

        ('Control de Asistencia', {
        'classes': ('collapse', 'tab', 'empty-form'),
        'fields': ('calendario_asistencia',),
        }),
        
    )

    # 2. READONLY_FIELDS: INCLUIMOS EL NUEVO CAMPO
    readonly_fields = ('mostrar_foto_de_base_de_datos', 'historial_posiciones_detalle','historico_movimientos_personal','calendario_asistencia',) # <-- CAMBIO AÑADIDO

    # -------------------------------------------------------------
    # --- NUEVO MÉTODO PARA GENERAR EL CRONOGRAMA DE POSICIONES ---
    # -------------------------------------------------------------
    @admin.display(description='Historial de Posiciones por Periodo')
    def historial_posiciones_detalle(self, obj):
        # El campo 'posicion' de PlantillaFin se enlaza con 'n_pos_actual' de MovPos
        posicion_actual = obj.posicion 
        
        # 1. Consulta la tabla MOV_POS (ordenado por fecha efectiva descendente)
        historial = MovPos.objects.filter(
            n_pos_actual=posicion_actual
        ).order_by('-f_efva') # f_efva es la Fecha Efectiva
        
        if not historial:
            return mark_safe("No se encontró historial para la posición actual.")
            
        # 2. Generación del HTML del cronograma (TIMELINE)
        html = f"""
        <style>
            .timeline-container {{ position: relative; padding: 20px 0; }}
            .timeline-item {{ border-left: 2px solid #ccc; padding-left: 20px; margin-bottom: 20px; position: relative; }}
            .timeline-dot {{ position: absolute; left: -10px; top: 0; width: 18px; height: 18px; border-radius: 50%; background-color: #337ab7; border: 3px solid white; }}
            .timeline-date {{ font-size: 0.9em; color: #777; margin-top: 0; }}
            .timeline-title {{ font-weight: bold; margin: 0 0 5px 0; color: #333; }}
            .timeline-current .timeline-dot {{ background-color: #5cb85c; }} /* Verde para la posición actual */
        </style>
        <div class="timeline-container">
        """
        
        for i, movimiento in enumerate(historial):
            es_actual = (i == 0) # Asumimos que el movimiento más reciente es el actual
            clase_css = 'timeline-current' if es_actual else ''
            
            # Usamos los campos relevantes de MOV_POS
            html += f"""
                <div class="timeline-item {clase_css}">
                    <div class="timeline-dot"></div>
                    <p class="timeline-date">
                        {movimiento.f_efva} (Captura: {movimiento.fecha_captura})
                    </p>
                    <h4 class="timeline-title">{movimiento.estado_psn}: {movimiento.motivo}</h4>
                    <p>Descripción: {movimiento.descr} / Ubicación: {movimiento.ubicacion}</p>
                </div>
            """
            
        html += """
        </div>
        """
        
        return mark_safe(html)
    
    @admin.display(description='Histórico de Movimientos de Personal')
    def historico_movimientos_personal(self, obj):

        if not obj:
            return ""

        # 🔑 Clave de enlace desde PlantillaFin
        num_empleado = getattr(obj, "num_empleado", None)

        if not num_empleado:
            return mark_safe("No se encontró número de empleado para enlazar movimientos en MOV_TOTAL.")

        # Consulta a tabla MOV_TOTAL a través del modelo Movimiento
        movimientos = (
            Movimiento.objects
            .filter(id_empl=str(num_empleado))
            .order_by('-fecha_efectiva','-sec','-fecha_captura')
        )

        if not movimientos:
            return mark_safe("No se encontraron movimientos de personal para este empleado en MOV_TOTAL.")

        html = """
        <style>
            .timeline-mov-container {
                padding: 10px 0;
            }
            /* FILA HORIZONTAL DE TARJETAS */
            .timeline-mov-row {
                display: flex;
                flex-wrap: wrap;          /* 👉 clave para que brinque a otra fila */
                gap: 10px;
                align-items: stretch;
            }
            /* TARJETA DE CADA MOVIMIENTO */
            .timeline-mov-card {
                position: relative;
                flex: 0 0 260px;          /* ancho aprox. de cada tarjeta */
                border: 1px solid #ccc;
                border-radius: 6px;
                padding: 8px 10px;
                background-color: #fdfdfd;
                box-shadow: 0 1px 2px rgba(0,0,0,0.05);
                font-size: 11px;
            }
            /* TARJETA DEL MOVIMIENTO MÁS RECIENTE */
            .timeline-mov-card.timeline-mov-current {
                border-color: #5cb85c;
                box-shadow: 0 0 0 2px rgba(92,184,92,0.3);
            }
            /* ETIQUETA "Actual" */
            .timeline-mov-badge {
                position: absolute;
                top: -8px;
                left: 10px;
                font-size: 10px;
                background: #800000;
                color: #fff;
                padding: 1px 6px;
                border-radius: 10px;
            }
            .timeline-mov-date {
                font-weight: bold;
                margin: 0 0 4px 0;
                color: #333;
            }
            .timeline-mov-title {
                margin: 0 0 3px 0;
                font-weight: bold;
                color: #444;
            }
            .timeline-mov-body {
                margin: 0;
                color: #555;
            }
        </style>
        <div class="timeline-mov-container">
            <div class="timeline-mov-row">
        """

        for i, mov in enumerate(movimientos):
            es_actual = (i == 0)  # Tomamos el movimiento más reciente como "último"
            clases = "timeline-mov-card"
            if es_actual:
                clases += " timeline-mov-current"

            html += f"""
                <div class="{clases}">
                    {'<span class="timeline-mov-badge">Actual</span>' if es_actual else ''}
                    <p class="timeline-mov-date">
                        {(mov.fecha_efectiva or '')[:10]} <span style="font-weight:normal;">(Captura: {(mov.fecha_captura or '')[:10]})</span>
                    </p>
                    <p class="timeline-mov-title">
                        Acción: {mov.accion_nombre or ''} &mdash; Motivo: {mov.motivo_nombre or ''}
                    </p>
                    <p class="timeline-mov-body">
                        <strong>Secuencia:</strong> {mov.sec or ''}<br/>                    
                        <strong>Estado de pago:</strong> {mov.estado_pago or ''}<br/>
                        <strong>Unidad:</strong> {mov.descr_larga1 or ''} /
                        <strong>Posición:</strong> {mov.posicion or ''}<br/>
                        <strong>Nivel:</strong> {mov.puesto_ptal or ''}<br/>
                        <strong>Puesto:</strong> {mov.descr_larga or ''}
                    </p>
                </div>
            """

        html += """
            </div>
        </div>
        """

        return mark_safe(html)

    # --- MÉTODOS ANTERIORES (SE MANTIENEN IGUAL) ---
    def _get_row_color(self, obj):
        if obj.programa == 'Sin Código' or obj.estado_nomina == 'Baja': return '#FFDDDD'
        elif obj.estado_nomina == 'Vacante' and 'H00' in str(obj.programa): return '#D4EDDA'
        return ''
    
    @admin.display(description='Foto')
    def colored_miniatura(self, obj):
        color = self._get_row_color(obj)
        content = self.mostrar_miniatura(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:1px; text-align:center;">{}</div>', color, content)

    @admin.display(description='Nombre Completo', ordering='nombres')
    def colored_nombres(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.nombres)

    @admin.display(description='RFC', ordering='rfc')
    def colored_rfc(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.rfc)

    @admin.display(description='Num Empleado', ordering='num_empleado')
    def colored_num_empleado(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.num_empleado)

    @admin.display(description='Posición', ordering='posicion')
    def colored_posicion(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.posicion)

    @admin.display(description='Nivel', ordering='nivel')
    def colored_nivel(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.nivel)

    @admin.display(description='Programa', ordering='programa')
    def colored_programa(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.programa)
        
    @admin.display(description='DG o Aduana', ordering='dg_o_aduana_compactada')
    def colored_dg_o_aduana_compactada(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.dg_o_aduana_compactada)

    @admin.display(description='Unidad Admin.', ordering='unidad_administrativa')
    def colored_unidad_administrativa(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.unidad_administrativa)

    @admin.display(description='Departamento', ordering='departamento')
    def colored_departamento(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.departamento)

    @admin.display(description='Puesto Funcional', ordering='nombre_puesto_funcional')
    def colored_nombre_puesto_funcional(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.nombre_puesto_funcional)

    @admin.display(description='Estado Nómina', ordering='estado_nomina')
    def colored_estado_nomina(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.estado_nomina)

    @admin.display(description='Tipo de personal SEDENA / SEMAR', ordering='tipo_de_personal_sedena_semar')
    def colored_tipo_de_personal_sedena_semar(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.tipo_de_personal_sedena_semar)

    @admin.display(description='Rango', ordering='rango')
    def colored_rango(self, obj):
        color = self._get_row_color(obj)
        return format_html('<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>', color, obj.rango)

    @admin.display(description='SMN', ordering='smn')
    def colored_smn(self, obj):
        color = self._get_row_color(obj)
        try:
            # Corrección: Usar la coma para miles en formato local, no el punto
            formatted_value = f"${obj.smn:,.2f}" 
        except (ValueError, TypeError, AttributeError):
            formatted_value = obj.smn
        return format_html(
            '<div style="width:100%; height:100%; background-color:{}; padding:5px;">{}</div>',
            color,
            formatted_value
        )

    def mostrar_foto_de_base_de_datos(self, obj):
        if obj.foto:
            imagen_base64 = base64.b64encode(obj.foto).decode('utf-8')
            return format_html(f'<img src="data:image/jpeg;base64,{imagen_base64}" style="max-width: 200px; height: auto;" />')
        return "No hay imagen en la base de datos."
    mostrar_foto_de_base_de_datos.short_description = 'Fotografía'

    def mostrar_miniatura(self, obj):
        if obj.foto:
            imagen_base64 = base64.b64encode(obj.foto).decode('utf-8')
            return format_html(f'<img src="data:image/jpeg;base64,{imagen_base64}" style="width: 60px; height: auto;" />')
        return "N/A"

    # --- ACCIÓN DE EXPORTAR CON REGISTRO EN BITÁCORA (IGUAL) ---
    @admin.action(description="Exportar a Excel con Imágenes")
    def exportar_a_excel_con_imagenes(self, request, queryset):
        RegistroActividad.objects.create(usuario=request.user, tipo_accion='export', detalles=f"Exportó a Excel {queryset.count()} registros de Plantilla de Personal.")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Plantilla de Personal"
        headers = ['Posición', 'Nombre', 'RFC', 'SMN', 'Foto']
        sheet.append(headers)
        for registro in queryset:
            if registro.foto:
                try:
                    image_stream = BytesIO(registro.foto)
                    img = Image(image_stream)
                    proporcion = img.width / img.height
                    img.height = 80
                    img.width = 80 * proporcion
                except Exception: img = None
            else: img = None
            row_data = [registro.posicion, registro.nombres, registro.rfc, getattr(registro, 'smn', ''), '']
            sheet.append(row_data)
            if img:
                row_num = sheet.max_row
                sheet.add_image(img, f'E{row_num}')
                sheet.row_dimensions[row_num].height = 65
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="plantilla_personal.xlsx"'
        workbook.save(response)
        return response

    # --- REGISTRO DE CONSULTAS Y VISTAS (IGUAL) ---
    def changelist_view(self, request, extra_context=None):
        search_term = request.GET.get('q', '') 
        if search_term:
            log_user_view(request, f"Buscó '{search_term}' en Plantilla de Personal.", include_filters=False) 
        else:
            log_user_view(request, "Consultó la lista de Plantilla de Personal.", include_filters=True) 
        return super().changelist_view(request, extra_context)

    def change_view(self, request, object_id, form_url='', extra_context=None):
        try:
            obj = self.get_object(request, object_id)
            log_user_view(request, f"Consultó el detalle de Plantilla: {obj}", include_filters=False) 
        except Exception: 
            log_user_view(request, f"Intentó consultar detalle de Plantilla ID: {object_id}", include_filters=False)
        return super().change_view(request, object_id, form_url, extra_context)
        
    # --- GESTIÓN DE ACCIONES Y PERMISOS (IGUAL) ---
    def get_actions(self, request):
        actions = super().get_actions(request)
        if 'delete_selected' in actions: del actions['delete_selected']
        return actions

    def has_delete_permission(self, request, obj=None):
        return False
    
    # -------------------------------------------------------------
    # --- MÉTODO PARA GENERAR EL CRONOGRAMA DE POSICIONES (AQUÍ VA EL CÓDIGO) ---
    # -------------------------------------------------------------
    @admin.display(description='Historial de Posiciones por Periodo')
    def historial_posiciones_detalle(self, obj):
        # El campo 'posicion' de PlantillaFin se enlaza con 'n_pos_actual' de MovPos
        posicion_actual = obj.posicion 
        
        # 1. Consulta la tabla MOV_POS (ordenado por fecha efectiva descendente)
        historial = MovPos.objects.filter(
            n_pos_actual=posicion_actual
        ).order_by('-f_efva') # f_efva es la Fecha Efectiva
        
        if not historial:
            return mark_safe("No se encontró historial para la posición actual.")
            
        # 2. Generación del HTML del cronograma (TIMELINE)
        html = f"""
        <style>
            .timeline-container {{ position: relative; padding: 20px 0; }}
            .timeline-item {{ border-left: 2px solid #ccc; padding-left: 20px; margin-bottom: 20px; position: relative; }}
            .timeline-dot {{ position: absolute; left: -10px; top: 0; width: 18px; height: 18px; border-radius: 50%; background-color: #337ab7; border: 3px solid white; }}
            .timeline-date {{ font-size: 0.9em; color: #777; margin-top: 0; }}
            .timeline-title {{ font-weight: bold; margin: 0 0 5px 0; color: #333; }}
            .timeline-current .timeline-dot {{ background-color: #5cb85c; }} /* Verde para la posición actual */
        </style>
        <div class="timeline-container">
        """
        
        for i, movimiento in enumerate(historial):
            es_actual = (i == 0) # Asumimos que el movimiento más reciente es el actual
            clase_css = 'timeline-current' if es_actual else ''
            
            # Usamos los campos relevantes de MOV_POS
            html += f"""
                <div class="timeline-item {clase_css}">
                    <div class="timeline-dot"></div>
                    <p class="timeline-date">
                        {movimiento.f_efva} (Captura: {movimiento.fecha_captura})
                    </p>
                    <h4 class="timeline-title">{movimiento.estado_psn}: {movimiento.motivo} / Tipo: {movimiento.descr}   </h4>
                    <p>Descripción: {movimiento.descr} / Departamento: {movimiento.departamento} / Área: {movimiento.unidad_de_negocio} </p>
                    <p>Nivel: {movimiento.nivel_ptal} / Puesto Funcional: {movimiento.puesto} </p>
                </div>
            """
            
        html += """
        </div>
        """
        
        return mark_safe(html)

    @admin.display(description='Calendario de asistencia')
    def calendario_asistencia(self, obj):
        if not obj:
            return ""

        num_empleado = getattr(obj, "num_empleado", None)
        if not num_empleado:
            return mark_safe("No se encontró número de empleado para enlazar accesos.")

        # Normalizar empleado: con y sin ceros a la izquierda
        empleado_raw = str(num_empleado).strip()
        empleado_sin_ceros = empleado_raw.lstrip('0') or empleado_raw

        # 1) Traer los accesos del empleado (de cualquier año)
        accesos = Acceso2025.objects.filter(
            Q(empleado=empleado_raw) | Q(empleado=empleado_sin_ceros)
        )

        if not accesos:
            return mark_safe("No se encontraron registros de accesos para este empleado.")

        # Helpers de nombres en español
        DIAS_ES = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        MESES_ES = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]

        # 2) Agrupar por fecha -> lista de (hora, temp)
        dias = defaultdict(list)  # { '2025-02-16': [ {'hora': '07:55:00', 'temp': 36.4, 'dt': datetime}, ... ] }
        fechas_dt = []  # para saber qué meses/años hay

        for acc in accesos:
            raw = acc.date_raw or ""
            if len(raw) < 19:
                continue

            fecha = raw[:10]         # 'YYYY-MM-DD'
            hora = raw[-8:]          # 'HH:MM:SS'

            try:
                dt = datetime.strptime(fecha, "%Y-%m-%d")
                fechas_dt.append(dt)
            except Exception:
                continue

            # Temperatura: REPLACE(left(param0,6),',','') pero tolerante
            temp_raw = ((acc.param0 or "")[:6]).strip()
            tmp = temp_raw.replace(',', '.')
            temp_clean = ''.join(ch for ch in tmp if (ch.isdigit() or ch in '.-'))

            if temp_clean:
                try:
                    temp_val = float(temp_clean)
                except Exception:
                    temp_val = temp_raw  # si no se puede a float, dejamos el string
            else:
                temp_val = None

            dias[fecha].append({
                "hora": hora,    # 'HH:MM:SS'
                "temp": temp_val,
                "dt": dt,
            })

        if not dias:
            return mark_safe("No se pudieron procesar registros de accesos válidos para este empleado.")

        # Helpers para promedios
        def hora_a_segundos(h):
            # h: 'HH:MM' o 'HH:MM:SS'
            if not h or h == "-":
                return None
            partes = h.split(":")
            if len(partes) < 2:
                return None
            try:
                hh = int(partes[0])
                mm = int(partes[1])
                ss = int(partes[2]) if len(partes) > 2 else 0
                return hh * 3600 + mm * 60 + ss
            except Exception:
                return None

        def segundos_a_hhmm(seg):
            if seg is None:
                return "-"
            hh = seg // 3600
            mm = (seg % 3600) // 60
            return f"{hh:02d}:{mm:02d}"

        # 3) Calcular entrada/salida por día y agrupar por mes
        meses = defaultdict(list)  # { '2025-02': [ { 'dt':..., 'fecha_label':..., 'entrada_hora':..., ... }, ... ] }

        for fecha_str, registros in dias.items():
            dt = registros[0]["dt"]  # todos los del día comparten fecha

            # separa en mañana (<12) y tarde (>=12)
            registros_manan = []
            registros_tarde = []

            for r in registros:
                hora_txt = r["hora"]
                try:
                    hh = int(hora_txt[:2])
                except Exception:
                    hh = 0
                if hh < 12:
                    registros_manan.append(r)
                else:
                    registros_tarde.append(r)

            # Entrada = menor hora de la mañana (si hay)
            entrada = min(registros_manan, key=lambda r: r["hora"]) if registros_manan else None
            # Salida = mayor hora de la tarde (si hay)
            salida = max(registros_tarde, key=lambda r: r["hora"]) if registros_tarde else None

            # Etiqueta tipo: "Lunes 3 Noviembre 2025"
            weekday_idx = dt.weekday()  # 0 = lunes
            dia_nombre = DIAS_ES[weekday_idx]
            mes_nombre = MESES_ES[dt.month - 1]
            etiqueta_dia = f"{dia_nombre} {dt.day} {mes_nombre} {dt.year}"

            clave_mes = dt.strftime("%Y-%m")     # '2025-11'
            nombre_mes = f"{mes_nombre} {dt.year}"

            # formato ‘HH:MM’ o '-'
            if entrada:
                entrada_hora_str = entrada["hora"][:5]
                entrada_temp_val = entrada["temp"]
            else:
                entrada_hora_str = "-"
                entrada_temp_val = None

            if salida:
                salida_hora_str = salida["hora"][:5]
                salida_temp_val = salida["temp"]
            else:
                salida_hora_str = "-"
                salida_temp_val = None

            meses[clave_mes].append({
                "dt": dt,
                "fecha_label": etiqueta_dia,
                "entrada_hora": entrada_hora_str,
                "entrada_temp": entrada_temp_val,
                "salida_hora": salida_hora_str,
                "salida_temp": salida_temp_val,
                "nombre_mes": nombre_mes,
            })

        # 4) Ordenar meses y decidir cuál es el "vigente" (mes actual si tiene datos)
        meses_ordenados = dict(sorted(meses.items(), key=lambda x: x[0]))  # orden cronológico

        hoy = timezone.localtime(timezone.now()).date()
        clave_mes_hoy = hoy.strftime("%Y-%m")  # ej. '2026-02'

        if clave_mes_hoy in meses_ordenados:
            vigente_clave_mes = clave_mes_hoy
        elif fechas_dt:
            dt_max = max(fechas_dt)
            vigente_clave_mes = dt_max.strftime("%Y-%m")
        else:
            vigente_clave_mes = next(iter(meses_ordenados.keys()))

        # 5) Construir HTML (nav + tablas por mes + estadísticas por mes)
        html = """
        <style>
            .asistencia-nav {
                margin: 8px 0 12px 0;
                font-size: 11px;
            }
            .asistencia-nav a {
                margin-right: 6px;
                text-decoration: none;
                padding: 3px 6px;
                border-radius: 3px;
                border: 1px solid #ccc;
                color: #333;
            }
            .asistencia-nav a.asistencia-nav-activo {
                background-color: #800000;
                color: #fff;
                border-color: #800000;
            }
            .asistencia-mes-title {
                margin-top: 10px;
                font-weight: bold;
                font-size: 13px;
                color: #800000;
            }
            .asistencia-tabla {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 10px;
                font-size: 11px;
            }
            .asistencia-tabla th, .asistencia-tabla td {
                border: 1px solid #ddd;
                padding: 4px 6px;
                text-align: center;
            }
            .asistencia-tabla th {
                background-color: #f5f5f5;
                font-weight: bold;
            }
            .asistencia-fila-par {
                background-color: #fafafa;
            }
            .asistencia-fila-impar {
                background-color: #ffffff;
            }
            .asistencia-mes-wrapper {
                display: none;
            }
            .asistencia-mes-wrapper.activo {
                display: block;
            }
            .asistencia-resumen-mes {
                font-size: 11px;
                margin: 4px 0 8px 0;
                padding: 4px 6px;
                background-color: #f0f0f0;
                border-radius: 4px;
            }
        </style>
        """

        # Navegación por meses (tabs)
        html += '<div class="asistencia-nav">'
        claves_mes_lista = list(meses_ordenados.keys())
        for clave_mes in claves_mes_lista:
            lista_dias = meses_ordenados[clave_mes]
            if not lista_dias:
                continue
            nombre_mes = lista_dias[0]["nombre_mes"]
            activo_cls = "asistencia-nav-activo" if clave_mes == vigente_clave_mes else ""
            html += f'<a href="#!" class="asistencia-nav-link {activo_cls}" data-mes="{clave_mes}">{nombre_mes}</a>'
        html += '</div>'

        # Contenedor de meses
        html += '<div id="asistencia-meses-contenedor">'

        for clave_mes, lista_dias in meses_ordenados.items():
            if not lista_dias:
                continue

            nombre_mes = lista_dias[0]["nombre_mes"]
            activo_cls = "activo" if clave_mes == vigente_clave_mes else ""

            # --- Estadísticas por mes ---
            entrada_segundos_mes = []
            salida_segundos_mes = []
            for d in lista_dias:
                ent_seg = hora_a_segundos(d["entrada_hora"])
                if ent_seg is not None:
                    entrada_segundos_mes.append(ent_seg)

                # solo contamos salida si realmente hay salida (no "-")
                if d["salida_hora"] and d["salida_hora"] != "-":
                    sal_seg = hora_a_segundos(d["salida_hora"])
                    if sal_seg is not None:
                        salida_segundos_mes.append(sal_seg)

            prom_entrada_mes = segundos_a_hhmm(
                sum(entrada_segundos_mes) // len(entrada_segundos_mes)
            ) if entrada_segundos_mes else "-"
            prom_salida_mes = segundos_a_hhmm(
                sum(salida_segundos_mes) // len(salida_segundos_mes)
            ) if salida_segundos_mes else "-"

            # --- Wrapper del mes ---
            html += f'<div class="asistencia-mes-wrapper {activo_cls}" id="asistencia-mes-{clave_mes}">'
            html += f'<div class="asistencia-mes-title">{nombre_mes}</div>'

            # Estadísticas mensuales encima de la tabla
            html += f"""
            <div class="asistencia-resumen-mes">
                <strong>Estadísticas del mes:</strong>
                Promedio de entrada: <strong>{prom_entrada_mes}</strong>,
                Promedio de salida: <strong>{prom_salida_mes}</strong>
            </div>
            """

            html += """
            <table class="asistencia-tabla">
                <thead>
                    <tr>
                        <th>Día</th>
                        <th>Hora Entrada</th>
                        <th>Temp. Entrada</th>
                        <th>Hora Salida</th>
                        <th>Temp. Salida</th>
                    </tr>
                </thead>
                <tbody>
            """

            # ordenar días del mes por fecha real
            lista_dias_ordenados = sorted(lista_dias, key=lambda d: d["dt"])

            for idx, dia in enumerate(lista_dias_ordenados):
                fila_clase = "asistencia-fila-par" if idx % 2 == 0 else "asistencia-fila-impar"

                # Formateo de temperaturas
                ent_temp_val = dia["entrada_temp"]
                sal_temp_val = dia["salida_temp"]

                if isinstance(ent_temp_val, (int, float, complex)):
                    temp_ent = f"{float(ent_temp_val):.1f}°"
                elif ent_temp_val:
                    temp_ent = f"{ent_temp_val}°"
                else:
                    temp_ent = "-"

                if isinstance(sal_temp_val, (int, float, complex)):
                    temp_sal = f"{float(sal_temp_val):.1f}°"
                elif sal_temp_val:
                    temp_sal = f"{sal_temp_val}°"
                else:
                    temp_sal = "-"

                html += f"""
                    <tr class="{fila_clase}">
                        <td>{dia['fecha_label']}</td>
                        <td>{dia['entrada_hora']}</td>
                        <td>{temp_ent}</td>
                        <td>{dia['salida_hora']}</td>
                        <td>{temp_sal}</td>
                    </tr>
                """

            html += """
                </tbody>
            </table>
            </div>
            """

        html += "</div>"

        # JS mínimo para cambiar de mes (y por tanto de estadísticas mensuales)
        html += """
        <script>
            (function() {
                var links = document.querySelectorAll('.asistencia-nav-link');
                if (!links.length) return;

                links.forEach(function(link) {
                    link.addEventListener('click', function(e) {
                        e.preventDefault();
                        var mes = this.getAttribute('data-mes');

                        // activar/desactivar links
                        links.forEach(function(l2) {
                            l2.classList.remove('asistencia-nav-activo');
                        });
                        this.classList.add('asistencia-nav-activo');

                        // mostrar/ocultar contenedores de meses
                        var wrappers = document.querySelectorAll('.asistencia-mes-wrapper');
                        wrappers.forEach(function(w) {
                            w.classList.remove('activo');
                        });
                        var activo = document.getElementById('asistencia-mes-' + mes);
                        if (activo) {
                            activo.classList.add('activo');
                        }
                    });
                });
            })();
        </script>
        """

        return mark_safe(html)
    
    @admin.display(description="Expediente")
    def expediente_link(self, obj):
        # No buscamos nada aquí. Solo mandamos al endpoint a demanda.
        url = reverse("myadmin:plantillafin_ultimo_expediente", args=[obj.pk])
        return format_html('<a class="exp-link" href="{}">Expediente</a>', url)

    @admin.display(description="Cédula")
    def generar_cedula_link(self, obj):
        # Si es un registro incompleto (como el verde), no mostrar botón
        if not obj.num_empleado or not obj.rfc or not obj.posicion:
            return ""

        url = reverse("generar_cedula_pdf", kwargs={"posicion_pk": obj.posicion})
        return format_html(
            '<a class="btn-cedula" href="{}" target="_blank">Cédula</a>',
            url
        )
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/ultimo-expediente/",
                self.admin_site.admin_view(self.ultimo_expediente_view),
                name="plantillafin_ultimo_expediente",
            ),
        ]
        return custom_urls + urls
    
    def ultimo_expediente_view(self, request, object_id):
        # 1) Obtén el empleado
        obj = self.get_object(request, object_id)
        if not obj:
            messages.error(request, "Empleado no encontrado.")
            return redirect("..")  # regresa al changelist

        # 2) Busca el último documento SOLO cuando se pide
        doc = (DocumentoExpediente.objects
            .filter(empleado_activo=obj)
            .exclude(archivo="")
            .order_by("-fecha_subida")
            .first())

        if not doc or not doc.archivo:
            messages.warning(request, "Este empleado no tiene expediente cargado.")
            return redirect(request.META.get("HTTP_REFERER", ".."))

        # 3) Redirige directo al PDF
        return redirect(doc.archivo.url)

# -------------------------------------------------------------
# --- CLASE DE ADMIN PARA BajasFin (SIN CAMBIOS) ---
# -------------------------------------------------------------
class BajasFinAdmin(admin.ModelAdmin):
    actions = ['exportar_a_excel_con_imagenes']
    
    inlines = [DocumentoExpedienteBajaInline]

    # ... (resto de la clase BajasFinAdmin, todo es correcto) ...
    list_display = ('mostrar_miniatura','nombre', 'rfc','id_empleado', 'posicion','nivel','motivo', 'fecha_efectiva_personal', 'puesto_funcional','unidad_administrativa')
    search_fields = ('nombre', 'id_empleado', 'rfc', 'motivo','puesto_funcional','nivel','unidad_administrativa')
    list_filter = (crear_filtro_conteo('motivo', 'Motivo'), crear_filtro_conteo('nivel', 'Nivel'), crear_filtro_conteo('unidad_administrativa', 'Unidad Administrativa'), crear_filtro_conteo('posicion', 'Posición'))
    fieldsets = (('Información del Empleado', {'fields': ('nombre', 'id_empleado', 'rfc', 'curp')}), ('Detalles de la Baja', {'fields': ('motivo', 'estado_nomina', 'fecha_efectiva_personal', 'fecha_de_captura')}), ('Información del Puesto', {'fields': ('posicion', 'puesto_funcional', 'unidad_administrativa')}), ('Fotografía', {'fields': ('mostrar_foto_de_base_de_datos',)}))
    readonly_fields = ('mostrar_foto_de_base_de_datos',)
    
    # --- MÉTODOS PARA FOTOS (Se mantienen igual) ---
    def mostrar_foto_de_base_de_datos(self, obj):
        if obj.foto:
            imagen_base64 = base64.b64encode(obj.foto).decode('utf-8')
            return format_html(f'<img src="data:image/jpeg;base64,{imagen_base64}" style="max-width: 200px; height: auto;" />')
        return "No hay imagen en la base de datos."
    mostrar_foto_de_base_de_datos.short_description = 'Fotografía'

    def mostrar_miniatura(self, obj):
        if obj.foto:
            imagen_base64 = base64.b64encode(obj.foto).decode('utf-8')
            return format_html(f'<img src="data:image/jpeg;base64,{imagen_base64}" style="width: 60px; height: auto;" />')
        return "N/A"
    mostrar_miniatura.short_description = 'Foto'

    # --- ACCIÓN DE EXPORTAR CON REGISTRO EN BITÁCORA (Se mantiene igual) ---
    @admin.action(description="Exportar Bajas seleccionadas a Excel con imágenes")
    def exportar_a_excel_con_imagenes(self, request, queryset):
        RegistroActividad.objects.create(usuario=request.user, tipo_accion='export', detalles=f"Exportó a Excel {queryset.count()} registros de Bajas de Personal.")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Bajas de Personal"
        headers = ['ID Empleado', 'Nombre', 'RFC', 'Posición', 'Motivo', 'Fecha Efectiva', 'Foto']
        sheet.append(headers)
        for registro in queryset:
            if registro.foto:
                try:
                    image_stream = BytesIO(registro.foto)
                    img = Image(image_stream)
                    proporcion = img.width / img.height
                    img.height = 80
                    img.width = 80 * proporcion
                except Exception: img = None
            else: img = None
            row_data = [registro.id_empleado, registro.nombre, registro.rfc, registro.posicion, registro.motivo, registro.fecha_efectiva_personal, '']
            sheet.append(row_data)
            if img:
                row_num = sheet.max_row
                sheet.add_image(img, f'G{row_num}')
                sheet.row_dimensions[row_num].height = 65
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="bajas_personal.xlsx"'
        workbook.save(response)
        return response

    # --- REGISTRO DE CONSULTAS Y VISTAS (Se mantiene igual) ---
    def changelist_view(self, request, extra_context=None):
        search_term = request.GET.get('q', '')
        if search_term:
            log_user_view(request, f"Buscó '{search_term}' en Bajas de Personal.", include_filters=False)
        else:
            log_user_view(request, "Consultó la lista de Bajas de Personal.", include_filters=True)
        return super().changelist_view(request, extra_context)

    def change_view(self, request, object_id, form_url='', extra_context=None):
        try:
            obj = self.get_object(request, object_id)
            log_user_view(request, f"Consultó el detalle de Baja: {obj}", include_filters=False)
        except Exception:
            log_user_view(request, f"Intentó consultar detalle de Baja ID: {object_id}", include_filters=False)
        return super().change_view(request, object_id, form_url, extra_context)

    # --- GESTIÓN DE ACCIONES (Se mantiene igual) ---
    def get_actions(self, request):
        actions = super().get_actions(request)
        if 'delete_selected' in actions: del actions['delete_selected']
        return actions

# -------------------------------------------------------------
# --- CLASE DE ADMIN PARA CalendarioNomina2025 (SIN CAMBIOS) ---
# -------------------------------------------------------------
class CalendarioNomina2025Admin(admin.ModelAdmin): # ¡Este es el nombre que faltaba!
    list_display = (
        'quincena',
        'id_periodo',
        'fecha_inicial',
        'fecha_final',
        # Incluye las columnas importantes para la visualización
        'aplicacion_movimientos_1_organizacion',
        'aplicacion_movimientos_2_organizacion',
        'aplicacion_movimientos_1_doaf',
        'aplicacion_movimientos_2_doaf',
        'fecha_procesamiento_nomina',
        'fecha_envio_a_finanzas',
        'fecha_envio_de_terceros',
        'fecha_envio_bancos',
        'fecha_pago',
        'clc_fonac',
        'fecha_de_pago_fonac',
    )
    # Define la estructura de la página de detalle
    fieldsets = (
        ('Periodo', { 'fields': ('quincena', 'id_periodo', 'fecha_inicial', 'fecha_final') }),
        ('Fechas de Movimientos y Nómina', {
            'fields': (
                'aplicacion_movimientos_1_organizacion',
                'aplicacion_movimientos_2_organizacion',
                'aplicacion_movimientos_1_doaf',
                'aplicacion_movimientos_2_doaf',
                'fecha_procesamiento_nomina',
                'fecha_envio_a_finanzas',
                'fecha_envio_de_terceros',
                'fecha_envio_bancos',
                'fecha_pago',
                'clc_fonac',
                'fecha_de_pago_fonac',
            )
        }),
    )
    # Los datos de esta tabla externa deben ser de solo lectura.
    readonly_fields = list_display 
    search_fields = ('quincena', 'id_periodo', 'fecha_inicial', 'fecha_final')
    list_filter = ('quincena',)

    # Se evita la edición, adición o borrado ya que es una tabla externa (managed=False)
    def has_add_permission(self, request):
        return False
    def has_change_permission(self, request, obj=None):
        return False
    def has_delete_permission(self, request, obj=None):
        return False

class DocumentoExpedienteAdmin(admin.ModelAdmin):
    list_display = (
        "num_empleado_text",
        "nombre_original",
        "tipo_documento",
        "anio_ingreso_anam",
        "empleado_activo",
        "empleado_baja",
        "estado_proceso",
        "fecha_subida",
    )
    search_fields = (
        "num_empleado_text",
        "nombre_original",
        "empleado_activo__nombres",
        "empleado_activo__rfc",
        "empleado_baja__nombre",
        "empleado_baja__rfc",
    )
    list_filter = ("estado_proceso", "anio_ingreso_anam", "tipo_documento")